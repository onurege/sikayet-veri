from flask import Flask, request, jsonify, send_file, render_template_string
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import time
import io
import openpyxl # Added
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import logging
import re # Added
import google.generativeai as genai # Added
from scraper import scrape_partners # Added

# Logging configuration
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# CORS configuration - production ready
CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:*", "http://127.0.0.1:*"],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})

# Ana sayfa - index.html'i serve et
@app.route('/')
def index():
    """Ana sayfayı göster"""
    try:
        # Eğer static/index.html varsa onu kullan
        if os.path.exists('static/index.html'):
            with open('static/index.html', 'r', encoding='utf-8') as f:
                return f.read()
        # Yoksa templates/index.html'i kullan
        elif os.path.exists('templates/index.html'):
            with open('templates/index.html', 'r', encoding='utf-8') as f:
                return f.read()
        # Hiçbiri yoksa basit bir sayfa göster
        else:
            return '''
            <html>
                <head><title>Şikayetvar Scraper</title></head>
                <body>    
                    <p>API Endpoints:</p>
                    <ul>
                        <li>/api/health - API durumu</li>
                        <li>/api/search?q=keyword - Arama</li>
                        <li>/api/export/excel - Excel export</li>
                    </ul>
                </body>
            </html>
            '''
    except Exception as e:
        logger.error(f"Index sayfası yüklenirken hata: {str(e)}")
        return jsonify({'error': 'Sayfa yüklenemedi'}), 500

@app.route('/api/search', methods=['GET'])
def search_complaints():
    """Şikayetvar.com'dan veri çeker - TÜM SAYFALAR"""
    
    keyword = request.args.get('q', '')
    fetch_all = request.args.get('all', 'true').lower() == 'true'
    search_type = request.args.get('type', 'company')
    
    if not keyword:
        return jsonify({'error': 'Anahtar kelime gerekli'}), 400
    
    all_complaints = []
    page = 1
    max_pages = 100
    total_pages_found = False
    consecutive_empty_pages = 0
    
    try:
        while page <= max_pages:
            print(f"\n🔄 Sayfa {page} çekiliyor...")
            
            # URL formatı
            if search_type == 'keyword' or '/' in keyword or ' ' in keyword:
                if page == 1:
                    url = f'https://www.sikayetvar.com/sikayetler?k={keyword}'
                else:
                    url = f'https://www.sikayetvar.com/sikayetler?k={keyword}&sayfa={page}'
            else:
                if page == 1:
                    url = f'https://www.sikayetvar.com/{keyword}'
                else:
                    url = f'https://www.sikayetvar.com/{keyword}?page={page}'
            
            print(f"📍 URL: {url}")
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'tr-TR,tr;q=0.9',
                'Connection': 'keep-alive',
                'Referer': f'https://www.sikayetvar.com/{keyword}'
            }
            
            session = requests.Session()
            response = session.get(url, headers=headers, timeout=15, allow_redirects=True)
            response.raise_for_status()
            
            # Yönlendirme kontrolü
            final_url = response.url
            expected_url = url.rstrip('/')
            actual_url = final_url.rstrip('/')
            
            if actual_url != expected_url:
                print(f" Sayfa {page} farklı URL'ye yönlendirildi:")
                print(f"   İstenilen: {expected_url}")
                print(f"   Gelen: {actual_url}")
                
                if page > 1 and '/sikayetler' in actual_url and '?k=' in actual_url:
                    if 'sayfa=' not in actual_url and 'page=' not in actual_url:
                        print(f" Son sayfaya ulaşıldı (ilk sayfaya yönlendirildi)")
                        break
                
                if page > 1 and actual_url == f'https://www.sikayetvar.com/{keyword}':
                    print(f" Son sayfaya ulaşıldı (ana sayfaya yönlendirildi)")
                    break
            
            if response.encoding is None or response.encoding == 'ISO-8859-1':
                response.encoding = 'utf-8'
            
            html = response.text
            
            if 'Ã' in html[:1000]:
                try:
                    html = response.content.decode('utf-8')
                except:
                    pass
            
            soup = BeautifulSoup(html, 'html.parser')
            
            complaint_cards = soup.find_all('article', class_='card-v2')
            
            if not complaint_cards:
                consecutive_empty_pages += 1
                print(f" Sayfa {page}'de şikayet kartı bulunamadı (Arka arkaya boş: {consecutive_empty_pages})")
                
                if consecutive_empty_pages >= 2:
                    print(f" 2 sayfa üst üste boş, aramayı sonlandırlıyor")
                    break
                
                page += 1
                import random
                wait_time = random.uniform(2, 4)
                print(f" {wait_time:.1f} saniye bekleniyor...")
                time.sleep(wait_time)
                continue
            
            print(f" Sayfa {page}'de {len(complaint_cards)} kart bulundu")
            
            page_complaints = []
            failed_cards = 0
            
            for idx, card in enumerate(complaint_cards):
                try:
                    # BAŞLIK
                    title = None
                    link = ''
                    
                    title_link = card.find('a', class_='complaint-description')
                    if not title_link:
                        title_link = card.find('a', class_='complaint-layer')
                    
                    if not title_link:
                        div_layer = card.find('div', class_='complaint-layer')
                        if div_layer:
                            title = div_layer.get_text(strip=True)
                            link = div_layer.get('data-complaint-link', '')
                            title_link = True
                    
                    if not title_link:
                        h2 = card.find('h2', class_='complaint-title')
                        if h2:
                            title_link = h2.find('a')
                            if not title_link:
                                div_layer = h2.find('div', class_='complaint-layer')
                                if div_layer:
                                    title = div_layer.get_text(strip=True)
                                    link = div_layer.get('data-complaint-link', '')
                                    title_link = True
                    
                    if title_link and isinstance(title_link, bool) == False:
                        title = title_link.get_text(strip=True)
                        link = title_link.get('href', '')
                    
                    if not title or len(title) < 5:
                        failed_cards += 1
                        continue
                    
                    if link and not link.startswith('http'):
                        link = 'https://www.sikayetvar.com' + link
                    
                    # ŞİRKET
                    company = 'Bilinmiyor'
                    company_elem = (
                        card.find('a', class_=lambda x: x and any(w in str(x).lower() for w in ['brand', 'company', 'firma'])) or
                        card.find('span', class_=lambda x: x and any(w in str(x).lower() for w in ['brand', 'company', 'firma']))
                    )
                    if company_elem:
                        company = company_elem.get_text(strip=True)
                    
                    # İÇERİK
                    content = ''
                    desc_elem = card.find('p', class_=lambda x: x and 'complaint-description' in str(x))
                    if not desc_elem:
                        desc_elem = card.find('div', class_=lambda x: x and 'complaint-description' in str(x))
                    if desc_elem:
                        content = desc_elem.get_text(strip=True)
                    
                    # DURUM
                    status = 'Beklemede'
                    status_elem = card.find('span', class_=lambda x: x and 'status' in str(x))
                    if status_elem:
                        status_text = status_elem.get_text(strip=True).lower()
                        if 'çözüldü' in status_text:
                            status = 'Çözüldü'
                        elif 'cevap' in status_text:
                            status = 'Cevaplandı'
                    
                    # UPVOTES
                    upvotes = 0
                    upvote_elem = card.find('span', class_=lambda x: x and 'rate-num' in str(x))
                    if upvote_elem:
                        try:
                            upvotes = int(''.join(filter(str.isdigit, upvote_elem.get_text())))
                        except:
                            pass
                    
                    # TARİH - Geliştirilmiş
                    date = ''
                    # 1. Deneme: time etiketi
                    date_elem = card.find('time')
                    if date_elem:
                        date = date_elem.get_text(strip=True)
                        if not date:
                            date = date_elem.get('title', '')
                            
                    # 2. Deneme: span içinde time class'ı
                    if not date:
                        date_elem = card.find('span', class_=lambda x: x and 'time' in str(x))
                        if date_elem:
                            date = date_elem.get_text(strip=True)
                            if not date:
                                date = date_elem.get('title', '')
                                
                    # 3. Deneme: div içinde time class'ı (Kullanıcı bildirimi: <div class="js-tooltip time tooltipstered">)
                    if not date:
                        date_elem = card.find('div', class_=lambda x: x and 'time' in str(x))
                        if date_elem:
                            date = date_elem.get_text(strip=True)
                            
                    # Tarih temizliği (Örn: "03 Eylül 2024 04:28400" -> "03 Eylül 2024 04:28")
                    if date:
                        # Genellikle tarih formatı "DD Ay YYYY HH:MM" şeklindedir
                        # Sondaki fazla rakamları kırpalım (muhtemelen view count vs karışıyor)
                        import re
                        # Sadece tarih ve saat formatını almaya çalışalım
                        # Örn: 03 Eylül 2024 09:00
                        match = re.search(r'(\d{1,2}\s+[a-zA-ZçÇğĞıİöÖşŞüÜ]+\s+\d{4}\s+\d{2}:\d{2})', date)
                        if match:
                            date = match.group(1)
                        else:
                            # Yıl yoksa (Örn: 04 Ağustos 14:02) -> Bu yılın tarihidir
                            # Grubu parçalayalım: (04 Ağustos) (14:02)
                            match_short = re.search(r'(\d{1,2}\s+[a-zA-ZçÇğĞıİöÖşŞüÜ]+)\s+(\d{2}:\d{2})', date)
                            if match_short:
                                current_year = datetime.now().year
                                # Araya yılı ekle: "04 Ağustos" + " 2025 " + "14:02"
                                date = f"{match_short.group(1)} {current_year} {match_short.group(2)}"

                    # 4. Deneme: kartın sağ üstündeki herhangi bir metadata
                    if not date:
                        meta_elem = card.find('div', class_='complaint-layer')
                        if meta_elem:
                             pass
                    
                    # ŞİKAYET ID
                    complaint_id = card.get('data-id', '') or card.get('id', '')
                    
                    # Veriyi ekle
                    complaint_data = {
                        'id': len(all_complaints) + idx + 1,
                        'complaint_id': complaint_id,
                        'title': title,
                        'company': company,
                        'content': content,
                        'status': status,
                        'upvotes': upvotes,
                        'date': date,
                        'link': link
                    }
                    
                    page_complaints.append(complaint_data)
                    
                except Exception as e:
                    failed_cards += 1
                    print(f"        Kart #{idx+1} parse hatası: {str(e)[:100]}")
                    continue
            
            if failed_cards > 0:
                print(f"   {failed_cards} kart parse edilemedi")
            
            if page_complaints:
                all_complaints.extend(page_complaints)
                print(f"   {len(page_complaints)} şikayet eklendi (Toplam: {len(all_complaints)})")
            
            if not fetch_all:
                print("   Tek sayfa modu, durduruldu")
                break
            
            if page >= 10 and len(all_complaints) > 200:
                print(f" Yeterli veri toplandı ({len(all_complaints)} şikayet), durduruldu")
                break
            
            page += 1
            
            # Rate limiting
            import random
            wait_time = random.uniform(1, 3)
            print(f"⏳ {wait_time:.1f} saniye bekleniyor...")
            time.sleep(wait_time)
        
        # Başarılı sonuç
        print(f"\n TAMAMLANDI: Toplam {len(all_complaints)} şikayet çekildi")
        
        return jsonify({
            'success': True,
            'data': all_complaints,
            'total': len(all_complaints),
            'keyword': keyword,
            'pages_scraped': page - 1,
            'timestamp': datetime.now().isoformat()
        })
        
    except requests.exceptions.RequestException as e:
        error_msg = str(e)
        print(f" İstek hatası: {error_msg}")
        
        if 'ConnectTimeout' in error_msg or 'ConnectionError' in error_msg:
            return jsonify({
                'error': 'Bağlantı zaman aşımı. Lütfen tekrar deneyin.',
                'success': False,
                'partial_data': all_complaints if all_complaints else None,
                'pages_scraped': page - 1
            }), 504
        
        if '429' in error_msg:
            return jsonify({
                'error': 'Çok fazla istek. Lütfen biraz bekleyin.',
                'success': False,
                'partial_data': all_complaints if all_complaints else None,
                'pages_scraped': page - 1
            }), 429
        
        return jsonify({
            'error': f'İstek hatası: {error_msg}',
            'success': False,
            'partial_data': all_complaints if all_complaints else None,
            'pages_scraped': page - 1
        }), 500
    except Exception as e:
        print(f" Beklenmeyen hata: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Beklenmeyen hata: {str(e)}',
            'success': False
        }), 500

@app.route('/api/export/excel', methods=['POST', 'OPTIONS'])
def export_excel():
    """Şikayetleri Excel formatında indir"""
    
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
        return response, 200
    
    try:
        data = request.json
        complaints = data.get('complaints', [])
        keyword = data.get('keyword', 'arama')
        
        if not complaints:
            return jsonify({'error': 'Veri bulunamadı'}), 400
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Şikayetler"
        
        headers = ['#', 'Şikayet ID', 'Başlık', 'Şirket', 'Durum', 'Upvote', 'Tarih', 'İçerik', 'Link']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, complaint in enumerate(complaints, 2):
            ws.cell(row=row_num, column=1, value=complaint.get('id', row_num-1))
            ws.cell(row=row_num, column=2, value=complaint.get('complaint_id', ''))
            
            title_cell = ws.cell(row=row_num, column=3, value=complaint.get('title', ''))
            title_cell.font = Font(bold=True, size=11)
            title_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            company_cell = ws.cell(row=row_num, column=4, value=complaint.get('company', ''))
            company_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            status = complaint.get('status', 'Beklemede')
            status_cell = ws.cell(row=row_num, column=5, value=status)
            status_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if status == 'Çözüldü':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
            elif status == 'Cevaplandı':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(color="9C6500")
            
            upvote_cell = ws.cell(row=row_num, column=6, value=complaint.get('upvotes', 0))
            upvote_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            date_cell = ws.cell(row=row_num, column=7, value=complaint.get('date', ''))
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            content_cell = ws.cell(row=row_num, column=8, value=complaint.get('content', ''))
            content_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            link = complaint.get('link', '')
            link_cell = ws.cell(row=row_num, column=9, value=link)
            if link:
                link_cell.hyperlink = link
                link_cell.font = Font(color="0563C1", underline="single")
            link_cell.alignment = Alignment(vertical='center')
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 70
        ws.column_dimensions['I'].width = 15
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(complaints)+1, min_col=1, max_col=9):
            for cell in row:
                cell.border = thin_border
        
        ws.freeze_panes = 'A2'
        
        summary_ws = wb.create_sheet(title="Özet")
        summary_ws['A1'] = 'Kurumsal Şikayet Analiz Platformu'
        summary_ws['A2'] = 'Pazar Analizi ve Müşteri Geri Bildirim Sistemi'
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws['A2'].font = Font(size=12, italic=True, color="666666")
        
        summary_data = [
            ['Arama Kelimesi:', keyword],
            ['Toplam Şikayet:', len(complaints)],
            ['Rapor Tarihi:', datetime.now().strftime('%d.%m.%Y %H:%M')],
            ['', ''],
            ['Durum Dağılımı:', ''],
        ]
        
        status_count = {}
        for c in complaints:
            status = c.get('status', 'Beklemede')
            status_count[status] = status_count.get(status, 0) + 1
        
        for status, count in status_count.items():
            summary_data.append([f'  {status}:', count])
        
        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_ws[f'A{row_num}'] = label
            summary_ws[f'B{row_num}'] = value
            summary_ws[f'A{row_num}'].font = Font(bold=True)
        
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 20
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f'sikayetvar_{keyword}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f" Excel export hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Excel oluşturma hatası: {str(e)}'}), 500

        return jsonify({'error': 'Sunucu hatası'}), 500

# AI Configuration

# Manual .env loader (Basit .env okuyucu)
try:
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()
        print(" .env dosyası başarıyla yüklendi.")
except Exception as e:
    print(f" .env yüklenirken hata: {e}")



@app.route('/api/analyze', methods=['POST'])
def analyze_complaints():
    """Gemini AI ile şikayetleri analiz et"""
    try:
        api_key = os.environ.get('GEMINI_API_KEY')
        if not api_key:
            return jsonify({
                'success': False, 
                'error': 'API Anahtarı bulunamadı. Lütfen GEMINI_API_KEY çevre değişkenini tanımlayın.'
            }), 400

        data = request.json
        complaints = data.get('complaints', [])
        
        if not complaints:
            return jsonify({'success': False, 'error': 'Analiz edilecek veri yok.'}), 400

        # Veriyi özetle (Token limitine takılmamak için sadece başlıkları ve kısa içerikleri alalım)
        # Son 50 şikayeti veya random örnekleri alabiliriz, şimdilik ilk 30 tanesini alalım detaylı analiz için.
        summary_text = "Müşteri Şikayetleri Listesi:\n"
        for idx, c in enumerate(complaints[:30]):
            summary_text += f"{idx+1}. Başlık: {c.get('title', '')} - Kurum: {c.get('company', '')} - İçerik: {c.get('content', '')[:100]}...\n"

        prompt = f"""
        Sen profesyonel bir veri analistisin. Aşağıdaki müşteri şikayetleri listesini analiz ederek Türkçe bir yönetici özeti çıkar. Rapora başlarken direkt olarak içerik vererek başla herhangi bir giriş cümlesi verme.
        
        Aşağıdaki formatta bir rapor sun:
        
        ##  Genel Duygu ve Durum
        (Kısa bir paragraf ile genel müşteri memnuniyetsizliği seviyesini ve tonunu özetle.)

        ##  Kök Nedenler
        (kaç tane şikayet gelmiş ve konu kirilimlari kategorileri kaçar tane, Bu konu Kirilimlarina göre özet bilgi ver, ve bu konu Kirilimlarina göre kök neden önerileri cikar)

        ## Analiz ve Aksiyon Önerileri
        (kök neden analizini ve aksiyon planını çalışmış oldukları dağıtım ve ERP sistemi ve depo yönetimi olarak hangi nedenlerde olduğunu analiz ederek yaz ve ona göre öneriler sun)
      
        ##  Öne Çıkan 3 Temel Sorun
        1. **[Sorun Başlığı]**: [Açıklama]
        2. **[Sorun Başlığı]**: [Açıklama]
        3. **[Sorun Başlığı]**: [Açıklama]

        ##  Aksiyon Önerileri
        * [Öneri 1]
        * [Öneri 2]
        * [Öneri 3]

        VERİLER:
        {summary_text}
        """

        genai.configure(api_key=api_key)
        
        # Model seçimi (Dinamik)
        target_model = 'gemini-pro' # Varsayılan
        try:
            # Kullanılabilir modelleri listele
            available_models = []
            for m in genai.list_models():
                if 'generateContent' in m.supported_generation_methods:
                    available_models.append(m.name)
            
            print(f"Kullanılabilir modeller: {available_models}")
            
            # Öncelik sırası: Flash > Pro > Diğerleri
            flash_model = next((m for m in available_models if 'flash' in m), None)
            pro_model = next((m for m in available_models if 'pro' in m and 'vision' not in m), None)
            
            if flash_model:
                target_model = flash_model
            elif pro_model:
                target_model = pro_model
            elif available_models:
                target_model = available_models[0]
                
            print(f"Seçilen model: {target_model}")
            
        except Exception as e:
            print(f"Model listeleme hatası: {e}, varsayılan kullanılıyor.")

        model = genai.GenerativeModel(target_model)
        
        response = model.generate_content(prompt)
        
        return jsonify({
            'success': True,
            'analysis': response.text
        })

    except Exception as e:
        logger.error(f"AI Analiz Hatası: {str(e)}")
        return jsonify({
            'success': False, 
            'error': f'AI Analizi sırasında hata oluştu: {str(e)}'
        }), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """API sağlık kontrolü"""
    return jsonify({
        'status': 'healthy',
        'message': 'Şikayetvar Scraper API çalışıyor',
        'timestamp': datetime.now().isoformat(),
        'environment': os.environ.get('FLASK_ENV', 'development'),
        'version': '1.0.0'
    })

# Error handler
@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Endpoint bulunamadı'}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Sunucu hatası'}), 500

@app.route('/api/partners', methods=['GET'])
def get_partners():
    """
    Endpoint to trigger partner scraping and return the data.
    """
    try:
        partners = scrape_partners()
        return jsonify(partners)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/partners/export/excel', methods=['POST'])
def export_partners_excel():
    """İş Ortaklarını Excel formatında indir"""
    try:
        data = request.json
        partners = data.get('partners', [])
        
        if not partners:
            return jsonify({'error': 'Dışa aktarılacak veri bulunamadı'}), 400
        
        wb = Workbook()
        ws = wb.active
        ws.title = "İş Ortakları"
        
        # Title Rows
        ws['A1'] = 'Kurumsal Şikayet Analiz Platformu'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center') 
        ws.merge_cells('A1:F1')
        
        ws['A2'] = 'Pazar Analizi ve Müşteri Geri Bildirim Sistemi'
        ws['A2'].font = Font(size=12, italic=True, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f'Rapor Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:F3')

        # Headers
        headers = ['#', 'Firma Adı', 'Şehir / İlçe', 'Telefon', 'E-posta', 'Web Sitesi']
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo color
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        start_row = 5 # Data starts after titles
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row_num, p in enumerate(partners, start_row + 1):
            # ID
            ws.cell(row=row_num, column=1, value=row_num-start_row).alignment = Alignment(horizontal='center')
            
            # Name
            name_cell = ws.cell(row=row_num, column=2, value=p.get('name', ''))
            name_cell.font = Font(bold=True)
            
            # Location
            ws.cell(row=row_num, column=3, value=p.get('location', ''))
            
            # Phone
            ws.cell(row=row_num, column=4, value=p.get('phone', ''))
            
            # Email
            email = p.get('email')
            email_cell = ws.cell(row=row_num, column=5, value=email if email else '')
            if email and '@' in email:
                 email_cell.hyperlink = f"mailto:{email}"
                 email_cell.font = Font(color="0563C1", underline="single")
            
            # Web
            web = p.get('web_address')
            web_cell = ws.cell(row=row_num, column=6, value=web if web else '')
            if web:
                if not web.startswith('http'): web = 'http://' + web
                web_cell.hyperlink = web
                web_cell.font = Font(color="0563C1", underline="single")
                
            # Borders
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = thin_border

        # Column Widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        
        ws.freeze_panes = 'A2'
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f'logo_is_ortaklari_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Partner Excel export hatası: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Port ayarı - environment variable veya default
    port = int(os.environ.get('PORT', 8000))
    
    # Debug modu - production'da False 
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    print("="*50)
    print("API start")
    print(f" Environment: {os.environ.get('FLASK_ENV', 'development')}")
    print(f" Debug Mode: {debug}")
    print("="*50)
    
    app.run(debug=debug, host='0.0.0.0', port=port)